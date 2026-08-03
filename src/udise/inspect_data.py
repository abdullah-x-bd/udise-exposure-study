from __future__ import annotations

import argparse
import csv
import fnmatch
import hashlib
import json
import os
import shutil
import sys
import tempfile
import zipfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import duckdb
import yaml
from charset_normalizer import from_bytes
from huggingface_hub import HfApi, hf_hub_download


SUPPORTED_TEXT = {".csv", ".tsv", ".txt"}
SUPPORTED_TABLES = SUPPORTED_TEXT | {".parquet", ".xlsx", ".xls"}


def load_yaml(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8") as handle:
        data = yaml.safe_load(handle)
    if not isinstance(data, dict):
        raise ValueError(f"{path} must contain a YAML mapping")
    return data


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(8 * 1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def human_bytes(value: int | None) -> str:
    if value is None:
        return "n/a"
    size = float(value)
    for unit in ("B", "KiB", "MiB", "GiB", "TiB"):
        if size < 1024 or unit == "TiB":
            return f"{size:.1f} {unit}"
        size /= 1024
    return f"{size:.1f} TiB"


def normalized_name(value: str) -> str:
    return value.replace("\ufeff", "").strip().lower()


def sql_literal(value: str) -> str:
    return "'" + value.replace("'", "''") + "'"


def quote_identifier(value: str) -> str:
    return '"' + value.replace('"', '""') + '"'


def detect_text_format(path: Path) -> tuple[str, str]:
    sample = path.read_bytes()[:256_000]
    match = from_bytes(sample).best()
    encoding = match.encoding if match and match.encoding else "utf-8"
    try:
        text = sample.decode(encoding, errors="replace")
    except LookupError:
        encoding = "utf-8"
        text = sample.decode("utf-8", errors="replace")

    suffix = path.suffix.lower()
    delimiter = "\t" if suffix == ".tsv" else ","
    try:
        dialect = csv.Sniffer().sniff(text, delimiters=",\t;|")
        delimiter = dialect.delimiter
    except csv.Error:
        pass
    return encoding, delimiter


def duckdb_encoding(encoding: str) -> str | None:
    value = encoding.lower().replace("_", "-")
    if value in {"ascii", "utf-8", "utf8", "utf-8-sig"}:
        return None
    aliases = {
        "cp1252": "windows-1252",
        "windows-1252": "windows-1252",
        "iso-8859-1": "latin-1",
        "latin-1": "latin-1",
        "latin1": "latin-1",
    }
    return aliases.get(value)


def csv_source_sql(
    path: Path,
    delimiter: str,
    encoding: str,
    sample_size: int,
    ignore_errors: bool,
) -> str:
    options = [
        "header=true",
        "all_varchar=true",
        f"sample_size={int(sample_size)}",
        f"ignore_errors={'true' if ignore_errors else 'false'}",
        "null_padding=true",
        f"delim={sql_literal(delimiter)}",
    ]
    encoded = duckdb_encoding(encoding)
    if encoded:
        options.append(f"encoding={sql_literal(encoded)}")
    return f"read_csv_auto({sql_literal(str(path))}, {', '.join(options)})"


def parquet_source_sql(path: Path) -> str:
    return f"read_parquet({sql_literal(str(path))})"


def column_lookup(columns: list[str]) -> dict[str, str]:
    return {normalized_name(column): column for column in columns}


def inspect_relation(
    connection: duckdb.DuckDBPyConnection,
    source_sql: str,
    table_key: str,
    expected_schema: dict[str, Any],
) -> dict[str, Any]:
    description = connection.execute(f"SELECT * FROM {source_sql} LIMIT 0").description
    columns = [item[0] for item in description]
    lookup = column_lookup(columns)

    expected = [normalized_name(item) for item in expected_schema.get("expected", [])]
    required = [normalized_name(item) for item in expected_schema.get("required", [])]
    actual = set(lookup)

    result: dict[str, Any] = {
        "columns": columns,
        "column_count": len(columns),
        "missing_expected_columns": sorted(set(expected) - actual),
        "unexpected_columns": sorted(actual - set(expected)),
        "missing_required_columns": sorted(set(required) - actual),
    }

    pseudocode = lookup.get("pseudocode")
    if pseudocode:
        identifier = quote_identifier(pseudocode)
        metrics = connection.execute(
            f"""
            WITH source AS (
                SELECT * FROM {source_sql}
            ),
            grouped AS (
                SELECT {identifier} AS pseudocode, COUNT(*) AS records
                FROM source
                WHERE {identifier} IS NOT NULL
                  AND TRIM(CAST({identifier} AS VARCHAR)) <> ''
                GROUP BY {identifier}
            )
            SELECT
                (SELECT COUNT(*) FROM source) AS row_count,
                COUNT(*) AS unique_pseudocode,
                COALESCE(SUM(records), 0) AS rows_with_pseudocode,
                COALESCE(SUM(CASE WHEN records > 1 THEN 1 ELSE 0 END), 0)
                    AS pseudocodes_with_multiple_rows,
                MIN(records) AS minimum_rows_per_school,
                MEDIAN(records) AS median_rows_per_school,
                MAX(records) AS maximum_rows_per_school,
                AVG(records) AS average_rows_per_school
            FROM grouped
            """
        ).fetchone()
        keys = [
            "row_count",
            "unique_pseudocode",
            "rows_with_pseudocode",
            "pseudocodes_with_multiple_rows",
            "minimum_rows_per_school",
            "median_rows_per_school",
            "maximum_rows_per_school",
            "average_rows_per_school",
        ]
        result["identifier_metrics"] = dict(zip(keys, metrics, strict=True))
        result["identifier_metrics"]["missing_pseudocode_rows"] = (
            result["identifier_metrics"]["row_count"]
            - result["identifier_metrics"]["rows_with_pseudocode"]
        )
    else:
        row_count = connection.execute(
            f"SELECT COUNT(*) FROM {source_sql}"
        ).fetchone()[0]
        result["identifier_metrics"] = {
            "row_count": row_count,
            "pseudocode_column_present": False,
        }

    item_group = lookup.get("item_group")
    item_id = lookup.get("item_id")
    if item_group and item_id and pseudocode:
        item_group_q = quote_identifier(item_group)
        item_id_q = quote_identifier(item_id)
        combinations = connection.execute(
            f"""
            SELECT
                CAST({item_group_q} AS VARCHAR) AS item_group,
                CAST({item_id_q} AS VARCHAR) AS item_id,
                COUNT(*) AS row_count,
                COUNT(DISTINCT {quote_identifier(pseudocode)})
                    AS school_count
            FROM {source_sql}
            GROUP BY 1, 2
            ORDER BY TRY_CAST(item_group AS INTEGER), TRY_CAST(item_id AS INTEGER),
                     item_group, item_id
            """
        ).fetchall()
        result["item_combinations"] = [
            {
                "item_group": row[0],
                "item_id": row[1],
                "row_count": row[2],
                "school_count": row[3],
            }
            for row in combinations
        ]

    return result


def inspect_delimited(
    path: Path,
    table_key: str,
    expected_schema: dict[str, Any],
    sample_size: int,
) -> dict[str, Any]:
    encoding, delimiter = detect_text_format(path)
    result: dict[str, Any] = {
        "format": "delimited_text",
        "detected_encoding": encoding,
        "detected_delimiter": repr(delimiter),
        "strict_parse_ok": True,
    }

    connection = duckdb.connect()
    try:
        source = csv_source_sql(
            path=path,
            delimiter=delimiter,
            encoding=encoding,
            sample_size=sample_size,
            ignore_errors=False,
        )
        try:
            result.update(
                inspect_relation(connection, source, table_key, expected_schema)
            )
        except Exception as strict_error:
            result["strict_parse_ok"] = False
            result["strict_parse_error"] = str(strict_error)
            fallback = csv_source_sql(
                path=path,
                delimiter=delimiter,
                encoding=encoding,
                sample_size=sample_size,
                ignore_errors=True,
            )
            result.update(
                inspect_relation(connection, fallback, table_key, expected_schema)
            )
            result["fallback_used"] = "ignore_errors=true"
    finally:
        connection.close()
    return result


def inspect_parquet(
    path: Path,
    table_key: str,
    expected_schema: dict[str, Any],
) -> dict[str, Any]:
    connection = duckdb.connect()
    try:
        result = {
            "format": "parquet",
            "strict_parse_ok": True,
        }
        result.update(
            inspect_relation(
                connection,
                parquet_source_sql(path),
                table_key,
                expected_schema,
            )
        )
        return result
    finally:
        connection.close()


def inspect_excel(path: Path, expected_schema: dict[str, Any]) -> dict[str, Any]:
    if path.suffix.lower() == ".xls":
        return {
            "format": "xls",
            "strict_parse_ok": False,
            "inspection_error": "Legacy XLS requires a separate conversion step.",
        }

    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheets: list[dict[str, Any]] = []
    expected = {
        normalized_name(item) for item in expected_schema.get("expected", [])
    }
    required = {
        normalized_name(item) for item in expected_schema.get("required", [])
    }

    for worksheet in workbook.worksheets:
        rows = worksheet.iter_rows(values_only=True)
        header = next(rows, ())
        columns = [str(value).strip() if value is not None else "" for value in header]
        actual = {normalized_name(value) for value in columns if value}
        sheets.append(
            {
                "sheet": worksheet.title,
                "reported_row_count": worksheet.max_row,
                "column_count": len(columns),
                "columns": columns,
                "missing_expected_columns": sorted(expected - actual),
                "unexpected_columns": sorted(actual - expected),
                "missing_required_columns": sorted(required - actual),
            }
        )
    workbook.close()
    return {
        "format": "xlsx",
        "strict_parse_ok": True,
        "sheets": sheets,
    }


def extract_member(
    archive: zipfile.ZipFile,
    member: zipfile.ZipInfo,
    destination: Path,
) -> Path:
    destination.mkdir(parents=True, exist_ok=True)
    target = destination / Path(member.filename).name
    with archive.open(member, "r") as source, target.open("wb") as output:
        shutil.copyfileobj(source, output, length=8 * 1024 * 1024)
    return target


def inspect_archive(
    archive_path: Path,
    table_key: str,
    remote_filename: str,
    expected_schema: dict[str, Any],
    sample_size: int,
) -> dict[str, Any]:
    result: dict[str, Any] = {
        "table": table_key,
        "remote_filename": remote_filename,
        "compressed_file_bytes": archive_path.stat().st_size,
        "sha256": sha256_file(archive_path),
        "archive_ok": True,
        "members": [],
        "data_members": [],
    }

    try:
        with zipfile.ZipFile(archive_path) as archive:
            bad_member = archive.testzip()
            result["zip_integrity_ok"] = bad_member is None
            if bad_member:
                result["first_bad_member"] = bad_member

            members = [item for item in archive.infolist() if not item.is_dir()]
            result["members"] = [
                {
                    "name": item.filename,
                    "compressed_bytes": item.compress_size,
                    "uncompressed_bytes": item.file_size,
                    "crc": item.CRC,
                }
                for item in members
            ]

            inspectable = [
                item
                for item in members
                if Path(item.filename).suffix.lower() in SUPPORTED_TABLES
            ]
            result["primary_member"] = (
                max(inspectable, key=lambda item: item.file_size).filename
                if inspectable
                else None
            )

            with tempfile.TemporaryDirectory(prefix=f"udise_{table_key}_") as temporary:
                temporary_path = Path(temporary)
                for member in inspectable:
                    extracted = extract_member(archive, member, temporary_path)
                    member_result: dict[str, Any] = {
                        "name": member.filename,
                        "uncompressed_bytes": member.file_size,
                    }
                    try:
                        suffix = extracted.suffix.lower()
                        if suffix in SUPPORTED_TEXT:
                            member_result.update(
                                inspect_delimited(
                                    extracted,
                                    table_key,
                                    expected_schema,
                                    sample_size,
                                )
                            )
                        elif suffix == ".parquet":
                            member_result.update(
                                inspect_parquet(
                                    extracted,
                                    table_key,
                                    expected_schema,
                                )
                            )
                        elif suffix in {".xlsx", ".xls"}:
                            member_result.update(
                                inspect_excel(extracted, expected_schema)
                            )
                    except Exception as error:
                        member_result.update(
                            {
                                "strict_parse_ok": False,
                                "inspection_error": str(error),
                            }
                        )
                    finally:
                        extracted.unlink(missing_ok=True)
                    result["data_members"].append(member_result)

            if not inspectable:
                result["archive_ok"] = False
                result["archive_error"] = "No supported tabular member found."
    except zipfile.BadZipFile as error:
        result["archive_ok"] = False
        result["archive_error"] = str(error)

    return result


def resolve_archives(
    repo_files: list[str],
    archive_config: dict[str, Any],
) -> dict[str, str]:
    resolved: dict[str, str] = {}
    errors: list[str] = []
    for table_key, settings in archive_config.items():
        pattern = settings["pattern"]
        matches = [
            filename
            for filename in repo_files
            if fnmatch.fnmatch(Path(filename).name, pattern)
        ]
        if len(matches) != 1:
            errors.append(
                f"{table_key}: pattern {pattern!r} matched {len(matches)} files: {matches}"
            )
        else:
            resolved[table_key] = matches[0]
    if errors:
        raise RuntimeError("\n".join(errors))
    return resolved


def build_markdown(manifest: dict[str, Any]) -> str:
    lines = [
        "# UDISE+ 2024-25 Source Inspection",
        "",
        f"Generated: {manifest['generated_at']}",
        "",
        f"Private dataset repository: `{manifest['dataset_repo']}`",
        "",
        "## Source archives",
        "",
        "| Table | Archive | Compressed | ZIP valid | Primary data member |",
        "|---|---|---:|:---:|---|",
    ]

    for archive in manifest["archives"]:
        lines.append(
            "| {table} | `{filename}` | {size} | {valid} | `{primary}` |".format(
                table=archive["table"],
                filename=archive["remote_filename"],
                size=human_bytes(archive.get("compressed_file_bytes")),
                valid="yes" if archive.get("zip_integrity_ok") else "no",
                primary=archive.get("primary_member") or "none",
            )
        )

    for archive in manifest["archives"]:
        lines.extend(["", f"## {archive['table']}", ""])
        if archive.get("archive_error"):
            lines.append(f"Archive error: `{archive['archive_error']}`")
            continue

        lines.append(
            f"Archive SHA-256: `{archive.get('sha256', 'n/a')}`"
        )
        lines.append("")
        lines.append(
            f"Archive members: {len(archive.get('members', []))}; "
            f"inspectable data members: {len(archive.get('data_members', []))}."
        )

        for member in archive.get("data_members", []):
            lines.extend(["", f"### `{member['name']}`", ""])
            lines.append(
                f"Format: `{member.get('format', 'unknown')}`; "
                f"uncompressed size: {human_bytes(member.get('uncompressed_bytes'))}."
            )
            if member.get("detected_encoding"):
                lines.append(
                    f"Detected encoding: `{member['detected_encoding']}`; "
                    f"delimiter: `{member.get('detected_delimiter')}`."
                )
            if not member.get("strict_parse_ok", False):
                lines.append(
                    f"Strict parsing did not succeed: "
                    f"`{member.get('strict_parse_error') or member.get('inspection_error')}`"
                )

            metrics = member.get("identifier_metrics")
            if metrics:
                lines.extend(["", "| Metric | Value |", "|---|---:|"])
                for key, value in metrics.items():
                    lines.append(f"| {key} | {value} |")

            columns = member.get("columns")
            if columns:
                lines.extend(
                    [
                        "",
                        f"Columns ({len(columns)}):",
                        "",
                        "```text",
                        "\n".join(columns),
                        "```",
                    ]
                )

            missing_required = member.get("missing_required_columns", [])
            missing_expected = member.get("missing_expected_columns", [])
            unexpected = member.get("unexpected_columns", [])
            if missing_required:
                lines.append(
                    f"Missing required columns: `{', '.join(missing_required)}`"
                )
            if missing_expected:
                lines.append(
                    f"Missing expected columns: `{', '.join(missing_expected)}`"
                )
            if unexpected:
                lines.append(
                    f"Unexpected columns: `{', '.join(unexpected)}`"
                )

            combinations = member.get("item_combinations", [])
            if combinations:
                lines.extend(
                    [
                        "",
                        "| item_group | item_id | rows | schools |",
                        "|---:|---:|---:|---:|",
                    ]
                )
                for item in combinations:
                    lines.append(
                        f"| {item['item_group']} | {item['item_id']} | "
                        f"{item['row_count']} | {item['school_count']} |"
                    )

    lines.extend(
        [
            "",
            "## Privacy",
            "",
            "This report contains structural metadata and aggregate counts only. "
            "It does not contain school-level records or sample rows.",
            "",
        ]
    )
    return "\n".join(lines)


def parse_arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--config",
        type=Path,
        default=Path("config/dataset.yml"),
    )
    parser.add_argument(
        "--schema",
        type=Path,
        default=Path("config/expected_schema.yml"),
    )
    return parser.parse_args()


def main() -> int:
    args = parse_arguments()
    config = load_yaml(args.config)
    expected_schemas = load_yaml(args.schema)

    repo_id = os.getenv(config["source"]["repo_env"], "").strip()
    token = os.getenv(config["source"]["token_env"], "").strip()
    if not repo_id:
        raise RuntimeError("HF_DATASET_REPO is not configured.")
    if not token:
        raise RuntimeError("HF_TOKEN is not configured.")

    output_dir = Path(config["inspection"]["output_dir"])
    output_dir.mkdir(parents=True, exist_ok=True)
    raw_dir = Path("data/raw")
    raw_dir.mkdir(parents=True, exist_ok=True)

    api = HfApi(token=token)
    repo_files = api.list_repo_files(repo_id=repo_id, repo_type="dataset")
    resolved = resolve_archives(repo_files, config["archives"])

    manifest: dict[str, Any] = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "dataset_repo": repo_id,
        "year": config["year"],
        "resolved_archives": resolved,
        "archives": [],
    }

    critical_errors: list[str] = []
    for table_key, remote_filename in resolved.items():
        local_path = Path(
            hf_hub_download(
                repo_id=repo_id,
                filename=remote_filename,
                repo_type="dataset",
                token=token,
                local_dir=raw_dir,
            )
        )
        archive_result = inspect_archive(
            archive_path=local_path,
            table_key=table_key,
            remote_filename=remote_filename,
            expected_schema=expected_schemas.get(table_key, {}),
            sample_size=int(config["inspection"]["csv_sample_size"]),
        )
        manifest["archives"].append(archive_result)
        if not archive_result.get("archive_ok"):
            critical_errors.append(
                f"{table_key}: {archive_result.get('archive_error', 'archive inspection failed')}"
            )
        local_path.unlink(missing_ok=True)

    manifest_path = output_dir / "source_manifest.json"
    report_path = output_dir / "inspection_report.md"
    manifest_path.write_text(
        json.dumps(manifest, indent=2, default=str),
        encoding="utf-8",
    )
    report = build_markdown(manifest)
    report_path.write_text(report, encoding="utf-8")

    summary_path = os.getenv("GITHUB_STEP_SUMMARY")
    if summary_path:
        with Path(summary_path).open("a", encoding="utf-8") as handle:
            handle.write(report)

    if critical_errors:
        print("\n".join(critical_errors), file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
